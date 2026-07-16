"""Phase 0.6 — sales spreadsheet inventory (2024-2026) + reconciliation vs Unleashed.

Counting rule (validated 13 Jul against three independent methods):
row = sale iff customer name present AND (Date col OR Week col parses as date)
AND dept not refund AND amount in (0, 40000) or blank. Kills the per-rep
scorecard blocks that sit below each ledger.
"""
import json
from datetime import datetime, date
from openpyxl import load_workbook
import google.auth
from google.cloud import bigquery
import pandas as pd

F25 = r'C:\Users\bashi\Downloads\2025 Sales Offline Version.xlsx'
F26 = r'C:\Users\bashi\Downloads\2026 Sales offline version.xlsx'
OUT = r'C:\Users\bashi\trustwarehouse\business_plan\analysis_output\data'

MONTH_NAMES = ["January", "February", "March", "April", "May", "June", "July",
               "August", "September", "October", "November", "December"]
TABS_25FILE = ([f"{m} 2024" for m in MONTH_NAMES] +
               [f"{m} 2025" for m in MONTH_NAMES[:7]])
TABS_26FILE = [f"{m} 2026" for m in MONTH_NAMES[:7]]

def month_key(tab):
    name, year = tab.rsplit(" ", 1)
    return "%s-%02d" % (year, MONTH_NAMES.index(name) + 1)

def is_date(v):
    if isinstance(v, (datetime, date)):
        return True
    if v is None:
        return False
    s = str(v).strip().split(" ")[0]
    for fmt in ("%d.%m.%y", "%d.%m.%Y", "%Y-%m-%d", "%d/%m/%y", "%d/%m/%Y"):
        try:
            datetime.strptime(s, fmt)
            return True
        except ValueError:
            continue
    return False

def count_tab(wb, tab):
    if tab not in wb.sheetnames:
        return None
    ws = wb[tab]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return None
    header = [str(c).strip().lower() if c else "" for c in rows[0]]
    def col(name, default=None):
        for i, h in enumerate(header):
            if h.startswith(name):
                return i
        return default
    ci_date, ci_name = col("date", 0), col("customer name", col("customer"))
    ci_amount, ci_dept, ci_week = col("amount"), col("dept"), col("week")
    if ci_name is None:
        return {"sales": None, "note": "no customer column: " + "|".join(h for h in header if h)[:80]}
    n, dom, rev = 0, 0, 0.0
    for r in rows[1:]:
        def cell(i):
            return r[i] if i is not None and i < len(r) else None
        name = cell(ci_name)
        if name is None or not str(name).strip():
            continue
        if not (is_date(cell(ci_date)) or is_date(cell(ci_week))):
            continue
        dept = str(cell(ci_dept) or "").strip()
        if "refund" in dept.lower():
            continue
        try:
            v = float(cell(ci_amount))
        except (TypeError, ValueError):
            v = None
        if v is not None and not (0 < v < 40000):
            continue
        n += 1
        if dept.lower() == "domestic":
            dom += 1
        if v:
            rev += v
    return {"sales": n, "domestic": dom, "revenue": round(rev, 2)}

results = {}
wb25 = load_workbook(F25, read_only=True, data_only=True)
for tab in TABS_25FILE:
    r = count_tab(wb25, tab)
    if r:
        results[month_key(tab)] = r
wb25.close()
wb26 = load_workbook(F26, read_only=True, data_only=True)
for tab in TABS_26FILE:
    r = count_tab(wb26, tab)
    if r:
        results[month_key(tab)] = r
wb26.close()

with open(OUT + r'\sheet_sales_by_month.json', 'w') as f:
    json.dump(results, f, indent=1)

# ---- Unleashed monthly for comparison ----
creds, _ = google.auth.default(scopes=['https://www.googleapis.com/auth/bigquery'])
client = bigquery.Client(project='trustwarehouse', credentials=creds)
un = client.query("""
  SELECT FORMAT_TIMESTAMP('%Y-%m',
           TIMESTAMP_MILLIS(SAFE_CAST(REGEXP_EXTRACT(order_date, r'([0-9]+)') AS INT64))) AS month,
         COUNT(*) AS un_orders,
         COUNTIF(order_status NOT IN ('Deleted','Parked')) AS un_orders_valid,
         ROUND(SUM(IF(order_status NOT IN ('Deleted','Parked'), SAFE_CAST(sub_total AS FLOAT64), 0)), 0) AS un_subtotal_exvat
  FROM `trustwarehouse.bronze.unleashed_sales_orders`
  GROUP BY month ORDER BY month
""").to_dataframe()

rows = []
months = sorted(set(list(results.keys()) + list(un.month)))
for m in months:
    sheet = results.get(m, {})
    u = un[un.month == m]
    rows.append({
        "month": m,
        "sheet_sales": sheet.get("sales"),
        "sheet_revenue": sheet.get("revenue"),
        "un_orders": int(u.un_orders_valid.iloc[0]) if len(u) else None,
        "un_subtotal_exvat": float(u.un_subtotal_exvat.iloc[0]) if len(u) else None,
    })
df = pd.DataFrame(rows)
df["count_delta"] = df.sheet_sales - df.un_orders
df.to_csv(OUT + r'\phase0_sheet_vs_unleashed.csv', index=False)
pd.set_option('display.width', 200)
print(df.to_string(index=False))

"""Phase 3B — rep scorecards.
Two populations, honestly separated:
- FIELD REPS (closers): from sales sheets (rep column) — sales, revenue, AOV,
  consistency, trend. No clean appointments-attended denominator exists
  (recorded nowhere structured) -> close-rate not computable; documented as gap.
- TELESALES AGENTS: unified calls (dials, conversations>=30s) + appointments
  booked (appointment_made_by picklist) monthly.
Rep-name variants resolved via explicit alias map. Fairness: min sample sizes;
lead-mix normalisation impossible for field reps (no assignment data) - stated.
"""
import json
from datetime import datetime, date
from openpyxl import load_workbook
import google.auth
from google.cloud import bigquery
import pandas as pd

creds, _ = google.auth.default(scopes=['https://www.googleapis.com/auth/bigquery'])
client = bigquery.Client(project='trustwarehouse', credentials=creds)
def q(sql):
    return client.query(sql).to_dataframe()
pd.set_option('display.width', 260)
OUT = r'C:\Users\bashi\trustwarehouse\business_plan\analysis_output\data'

REP_ALIASES = {
    "kris": "Kris", "kourosh": "Kris",
    "sam": "SamC", "samc": "SamC", "sam chapman": "SamC",
    "samantha": "Samantha", "sammy": "Samuel", "samuel": "Samuel",
    "chrisk": "ChrisK", "chris k": "ChrisK", "chris krammer": "ChrisK", "chris kramer": "ChrisK",
    "chrism": "ChrisM", "chris m": "ChrisM", "chris mannix": "ChrisM",
    "chriss": "ChrisS", "chris s": "ChrisS", "chris southworth": "ChrisS",
    "rob": "Rob", "rob chapman": "Rob", "kelly": "Kelly", "kelly miller": "Kelly",
    "niall": "Niall", "scott": "Scott", "josh": "Josh", "dec": "Dec",
    "paul": "Paul", "paul slade": "Paul", "office": "Office", "chc": "CHC",
}
def canon(rep):
    r = str(rep or "").strip().lower()
    return REP_ALIASES.get(r, str(rep or "").strip())

MONTH_NAMES = ["January","February","March","April","May","June","July",
               "August","September","October","November","December"]
def is_date(v):
    if isinstance(v, (datetime, date)): return True
    if v is None: return False
    s = str(v).strip().split(" ")[0]
    for fmt in ("%d.%m.%y","%d.%m.%Y","%Y-%m-%d","%d/%m/%y","%d/%m/%Y"):
        try: datetime.strptime(s, fmt); return True
        except ValueError: continue
    return False

rows_out = []
for path, tabs in [
    (r'C:\Users\bashi\Downloads\2025 Sales Offline Version.xlsx',
     [f"{m} 2024" for m in MONTH_NAMES[7:]] + [f"{m} 2025" for m in MONTH_NAMES[:7]]),
    (r'C:\Users\bashi\Downloads\2026 Sales offline version.xlsx',
     [f"{m} 2026" for m in MONTH_NAMES[:7]]),
]:
    wb = load_workbook(path, read_only=True, data_only=True)
    for tab in tabs:
        if tab not in wb.sheetnames: continue
        ws = wb[tab]
        rws = list(ws.iter_rows(values_only=True))
        header = [str(c).strip().lower() if c else "" for c in rws[0]]
        def col(name, default=None):
            for i, h in enumerate(header):
                if h.startswith(name): return i
            return default
        ci_d, ci_n = col("date", 0), col("customer name", col("customer"))
        ci_a, ci_dept, ci_w = col("amount"), col("dept"), col("week")
        ci_rep = col("rep", col("person"))
        name0, year = tab.rsplit(" ", 1)
        mk = "%s-%02d" % (year, MONTH_NAMES.index(name0) + 1)
        for r in rws[1:]:
            def cell(i): return r[i] if i is not None and i < len(r) else None
            nm = cell(ci_n)
            if nm is None or not str(nm).strip(): continue
            if not (is_date(cell(ci_d)) or is_date(cell(ci_w))): continue
            dept = str(cell(ci_dept) or "").strip()
            if "refund" in dept.lower(): continue
            try: v = float(cell(ci_a))
            except (TypeError, ValueError): v = None
            if v is not None and not (0 < v < 40000): continue
            rows_out.append({"month": mk, "rep": canon(cell(ci_rep)), "amount": v or 0.0})
    wb.close()

sheet = pd.DataFrame(rows_out)
sheet.to_csv(OUT + r'\phase3_sheet_sales_rows.csv', index=False)

print("=== FIELD REP scorecard — last 12 available sheet months (Jul25 + Jan-Jul26) ===")
recent = sheet[(sheet.month >= '2025-07')]
sc = recent.groupby('rep').agg(sales=('amount','count'), revenue=('amount','sum')).reset_index()
sc['aov'] = (sc.revenue / sc.sales).round(0)
sc = sc[sc.sales >= 10].sort_values('revenue', ascending=False)
sc['rev_share_pct'] = (sc.revenue / sc.revenue.sum() * 100).round(1)
print(sc.to_string(index=False))
print("top-2 revenue concentration: %.1f%%" % sc.rev_share_pct.head(2).sum())

print("\n=== field rep monthly consistency (2026 H1, sales/month) ===")
cons = sheet[sheet.month >= '2026-01'].groupby(['rep','month']).size().unstack(fill_value=0)
cons = cons[cons.sum(axis=1) >= 10]
print(cons.to_string())

print("\n=== TELESALES scorecard — last 12 months ===")
print(q("""
  WITH uc AS (
    SELECT LOWER(ANY_VALUE(direction)) AS direction,
           ARRAY_AGG(_colleague_name ORDER BY talk_time DESC LIMIT 1)[OFFSET(0)] AS agent,
           SUM(talk_time) AS talk, TIMESTAMP_MILLIS(MIN(start_time)) AS start
    FROM (SELECT DISTINCT id, flow_index, start_time, direction, _colleague_name, talk_time
          FROM `trustwarehouse.bronze.wildix_calls`)
    GROUP BY id HAVING start >= '2025-07-01' AND start < '2026-07-01'
    UNION ALL
    SELECT direction, JSON_VALUE(`from`, '$.name'), duration, start
    FROM `trustwarehouse.bronze.ascend_calls`
    WHERE start >= '2026-07-01' AND direction = 'outbound'),
  dials AS (
    SELECT agent, COUNT(*) AS dials, COUNTIF(talk >= 30) AS conversations,
           ROUND(SUM(talk)/3600, 0) AS talk_hours
    FROM uc WHERE direction = 'outbound' GROUP BY agent),
  booked AS (
    SELECT TRIM(appointment_made_by_65e1a90253305) AS agent, COUNT(*) AS appts_booked
    FROM `trustwarehouse.bronze.sharpspring_leads`
    WHERE create_timestamp >= '2025-07-01'
      AND TRIM(COALESCE(status_633ae6f6ac6fe,'')) IN
          ('Appointment','Appointment Cancelled','WhatsApp Appointment')
      AND TRIM(COALESCE(appointment_made_by_65e1a90253305,'')) != ''
    GROUP BY 1)
  SELECT d.agent, d.dials, d.conversations, d.talk_hours, b.appts_booked,
         ROUND(b.appts_booked / NULLIF(d.conversations,0) * 100, 1) AS appts_per_100_conv
  FROM dials d
  LEFT JOIN booked b ON LOWER(SPLIT(b.agent, ' ')[OFFSET(0)]) = LOWER(SPLIT(d.agent, ' ')[OFFSET(0)])
  WHERE d.dials >= 1000
  ORDER BY b.appts_booked DESC
""").to_string(index=False))

print("\n=== reconciliation: sheet sales per rep vs CRM 'Sold' owner, 2026 H1 ===")
crm = q("""
  SELECT owner_id, COUNT(*) AS crm_sold
  FROM `trustwarehouse.bronze.sharpspring_leads`
  WHERE TRIM(COALESCE(status_633ae6f6ac6fe,'')) = 'Sold'
    AND create_timestamp >= '2026-01-01'
  GROUP BY 1 ORDER BY crm_sold DESC LIMIT 8
""")
print(crm.to_string(index=False))
h1 = sheet[(sheet.month >= '2026-01') & (sheet.month <= '2026-06')].groupby('rep').size()
print("sheet 2026 H1 per rep:", h1.sort_values(ascending=False).head(10).to_dict())
